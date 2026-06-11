"""
히트펌프 경제성 분석 솔루션
─────────────────────────────────────────────────────────────
시민이 거주 환경과 에너지 사용량을 입력하면, 친환경 히트펌프(AWHP)
전환 시 경제적 이득과 환경 기여도를 계산해주는 Streamlit 앱입니다.

데이터 출처: 전기요금완료본.xlsx (요금제 × 태양광 × 난방유형 20개 블록)

[2026-06 수정] 누진제 HP 전기요금에 ① 슈퍼유저요금(동·하계 1,000kWh 초과분
736.2원/kWh)과 ② 증분비용 방식(HP가 일으킨 한계 비용만 HP에 귀속)을 반영.
앱 결과는 본 Python 코드에서 산출되므로(엑셀 블록은 결과 미사용), 두 효과 모두
여기 코드에 구현되어야 실제 HP 비용에 반영됨 — 이중계산 위험 없음.
"""
import os
import pandas as pd
import streamlit as st
import altair as alt
from PIL import Image
from openpyxl import load_workbook


st.set_page_config(page_title="GESI 히트펌프 경제성·환경성 간이 계산기", layout="wide")


# ══════════════════════════════════════════════════════════════════════
# 1. 스타일 정의
# ══════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');

:root {
    /* Surface / 배경 톤 */
    --bg-canvas:        #fafaf9;   /* 전체 배경 */
    --bg-surface:       #ffffff;   /* 카드 */
    --bg-soft:          #f5f5f4;   /* 옅은 회색 표면 */
    --bg-sage:          #f0fdf4;   /* 세이지 — saving-box 기본 */
    --bg-teal:          #e6f4f1;   /* 옅은 틸 — efficiency box */
    --bg-amber:         #fef7e6;   /* 옅은 골드 — 강조행 */

    /* Border */
    --border-subtle:    #e7e5e4;   /* stone-200 */
    --border-medium:    #d6d3d1;   /* stone-300 */
    --border-sage:      #a7d4be;   /* sage 200 */
    --border-teal:      #99d4c8;   /* teal 200 */

    /* Text */
    --text-primary:     #1c1917;   /* stone-900 */
    --text-secondary:   #44403c;   /* stone-700 */
    --text-muted:       #78716c;   /* stone-500 */

    /* Accent */
    --accent-emerald:   #047857;   /* 절감/긍정 */
    --accent-emerald-d: #064e3b;   /* 더 진한 강조 */
    --accent-teal:      #0f766e;   /* HP/대안 */
    --accent-red:       #dc2626;   /* 가스/기존 */
    --accent-slate:     #475569;   /* 메타데이터 배지 */

    /* Elevation */
    --shadow-sm:        0 1px 2px rgba(28,25,23,0.04), 0 1px 3px rgba(28,25,23,0.06);
    --shadow-md:        0 2px 4px rgba(28,25,23,0.04), 0 4px 12px rgba(28,25,23,0.08);
}

* { font-family: 'Pretendard', sans-serif; }

/* 박스·타이틀 */
.info-box      { background:var(--bg-surface); border:1px solid var(--border-subtle); border-radius:14px; padding:28px 32px; margin-bottom:36px; box-shadow:var(--shadow-sm); }
.info-title    { color:var(--text-primary); font-size:1.25rem; font-weight:700; margin-bottom:14px; margin-top:0; letter-spacing:-0.01em; }
.info-text     { color:var(--text-secondary); font-size:1.0rem; line-height:1.75; margin-bottom:0; }
.section-title { color:var(--text-primary); font-weight:700; font-size:1.3rem; margin-top:44px; margin-bottom:18px; border-bottom:1px solid var(--border-subtle); padding-bottom:10px; letter-spacing:-0.01em; }
.help-text     { color:var(--text-muted); font-size:0.85rem; margin-bottom:12px; line-height:1.5; }

/* 환경 기여 박스 (기본 = sage) */
.saving-box    { background:var(--bg-sage); border:1px solid var(--border-sage); border-radius:14px; padding:22px 26px; margin:16px 0; box-shadow:var(--shadow-sm); }
.saving-title  { color:var(--accent-emerald-d); font-size:1.1rem; font-weight:700; margin-bottom:6px; letter-spacing:-0.005em; }
.saving-sub    { color:var(--accent-emerald); font-size:0.95rem; line-height:1.7; }

/* 배지 (요금제·난방·규모 표시) — 차분한 슬레이트 톤 */
.tariff-badge  { display:inline-block; background:var(--bg-soft); color:var(--accent-slate); border:1px solid var(--border-subtle); padding:4px 11px; border-radius:8px; font-size:0.85rem; font-weight:600; margin-right:6px; }
.solar-badge-x { display:inline-block; background:var(--bg-amber); color:#854d0e; border:1px solid #fde68a; padding:4px 11px; border-radius:8px; font-size:0.85rem; font-weight:600; margin-right:6px; }
.solar-badge-o { display:inline-block; background:var(--bg-sage); color:var(--accent-emerald-d); border:1px solid var(--border-sage); padding:4px 11px; border-radius:8px; font-size:0.85rem; font-weight:600; margin-right:6px; }

/* 호버 툴팁 */
.has-tooltip { position:relative; cursor:help; }
.has-tooltip::after {
    content: attr(data-tooltip);
    position: absolute;
    bottom: calc(100% + 10px); left: 50%;
    transform: translateX(-50%);
    background: var(--text-primary); color: #ffffff;
    padding: 10px 14px; border-radius: 8px;
    font-size: 0.82rem; font-weight: 400;
    width: 300px; white-space: normal;
    line-height: 1.55; text-align: left;
    opacity: 0; visibility: hidden;
    pointer-events: none;
    transition: opacity 0.2s, visibility 0.2s;
    z-index: 1000;
    box-shadow: var(--shadow-md);
}
.has-tooltip::before {
    content: '';
    position: absolute;
    bottom: calc(100% + 4px); left: 50%;
    transform: translateX(-50%);
    border: 6px solid transparent;
    border-top-color: var(--text-primary);
    opacity: 0; visibility: hidden;
    pointer-events: none;
    transition: opacity 0.2s, visibility 0.2s;
    z-index: 1000;
}
.has-tooltip:hover::after,
.has-tooltip:hover::before { opacity: 1; visibility: visible; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
# 2. 정적 데이터 및 상수
# ══════════════════════════════════════════════════════════════════════

# ── 광역·기초 지자체 목록 ──
REGIONS_FULL = {
    "서울": ["종로구","중구","용산구","성동구","광진구","동대문구","중랑구","성북구","강북구","도봉구","노원구","은평구","서대문구","마포구","양천구","강서구","구로구","금천구","영등포구","동작구","관악구","서초구","강남구","송파구","강동구"],
    "강원도": ["춘천시","원주시","강릉시","동해시","태백시","속초시","삼척시","홍천군","횡성군","영월군","평창군","정선군","철원군","화천군","양구군","인제군","고성군","양양군"],
    "경기도": ["수원시","고양시","용인시","성남시","부천시","화성시","안산시","남양주시","안양시","평택시","시흥시","파주시","의정부시","김포시","광주시","광명시","군포시","하남시","오산시","양주시","이천시","구리시","안성시","포천시","의왕시","여주시","동두천시","과천시","가평군","양평군","연천군"],
    "제주도": ["제주시","서귀포시"],
    "인천": ["중구","동구","미추홀구","연수구","남동구","부평구","계양구","서구","강화군","옹진군"],
    "부산": ["중구","서구","동구","영도구","부산진구","동래구","남구","북구","해운대구","사하구","금정구","강서구","연제구","수영구","사상구","기장군"],
    "대구": ["중구","동구","서구","남구","북구","수성구","달서구","달성군"],
    "세종": ["세종특별자치시"],
    "대전": ["동구","중구","서구","유성구","대덕구"],
    "울산": ["중구","남구","동구","북구","울주군"],
    "광주": ["동구","서구","남구","북구","광산구"],
    "충청북도": ["청주시","충주시","제천시","보은군","옥천군","영동군","증평군","진천군","괴산군","음성군","단양군"],
    "충청남도": ["천안시","공주시","보령시","아산시","서산시","논산시","계룡시","당진시","금산군","부여군","서천군","청양군","홍성군","예산군","태안군"],
    "전라북도": ["전주시","군산시","익산시","정읍시","남원시","김제시","완주군","진안군","무주군","장수군","임실군","순창군","고창군","부안군"],
    "전라남도": ["목포시","여수시","순천시","나주시","광양시","담양군","곡성군","구례군","고흥군","보성군","화순군","장흥군","강진군","해남군","영암군","무안군","함평군","영광군","장성군","완도군","진도군","신안군"],
    "경상북도": ["포항시","경주시","김천시","안동시","구미시","영주시","영천시","상주시","문경시","경산시","의성군","청송군","영양군","영덕군","청도군","고령군","성주군","칠곡군","예천군","봉화군","울진군","울릉군"],
    "경상남도": ["창원시","진주시","통영시","사천시","김해시","밀양시","거제시","양산시","의령군","함안군","창녕군","고성군","남해군","하동군","산청군","함양군","거창군","합천군"],
}

# ── HDD(난방도일) 및 기후존별 sCOP는 엑셀의 'COP_계산기' 시트에서 동적 로드됩니다. ──
# Tbase=18°C 기준, 각 지역 월별 데이터는 load_tariff_xlsx()를 통해 가져옵니다.

# ── 엑셀 블록 헤더 위치 ──
# (요금제, 태양광플래그, 난방유형) → 헤더 행 번호
# 일반용은 별도 미터(HP 전용)이라 태양광 영향 없음 → 태X만 존재
EXCEL_BLOCK_HEADERS = {
    ("누진제", "태X", "도시가스(콘덴싱)"):   5, ("누진제", "태X", "도시가스(일반)"):  18,
    ("누진제", "태X", "등유"):              31, ("누진제", "태X", "LPG"):            44,
    ("누진제", "태O", "도시가스(콘덴싱)"):  57, ("누진제", "태O", "도시가스(일반)"):  70,
    ("누진제", "태O", "등유"):              83, ("누진제", "태O", "LPG"):            96,
    ("일반용", "태X", "도시가스(콘덴싱)"): 109, ("일반용", "태X", "도시가스(일반)"): 122,
    ("일반용", "태X", "등유"):             135, ("일반용", "태X", "LPG"):           148,
    ("계시별", "태X", "도시가스(콘덴싱)"): 161, ("계시별", "태X", "도시가스(일반)"): 174,
    ("계시별", "태X", "등유"):             187, ("계시별", "태X", "LPG"):           200,
    ("계시별", "태O", "도시가스(콘덴싱)"): 213, ("계시별", "태O", "도시가스(일반)"): 226,
    ("계시별", "태O", "등유"):             239, ("계시별", "태O", "LPG"):           252,
}

# ── UI 옵션 → 엑셀 키 매핑 ──
HEATING_TYPE_MAP = {
    "가스 콘덴싱 보일러": "도시가스(콘덴싱)",
    "가스 일반 보일러":   "도시가스(일반)",
    "등유 보일러":        "등유",
    "LPG 보일러":         "LPG",
}

# ── 5개 요금제 탭 → (엑셀 요금제, 엑셀 태양광 플래그) ──
# 사용자에게는 "태양광 설치/미설치"로 표기, 엑셀 내부 키는 태O/태X 그대로 유지
TARIFF_LABEL_MAP = {
    "누진제 (태양광 미설치)": ("누진제", "태X"),
    "누진제 (태양광 설치)":   ("누진제", "태O"),
    "일반용 (HP 전용 미터)":  ("일반용", "태X"),
    "계시별 (태양광 미설치)": ("계시별", "태X"),
    "계시별 (태양광 설치)":   ("계시별", "태O"),
}

# ── CO₂ 배출 환산 계수 ──
# 출처: 엑셀 Sheet2 행 10 "기기 온실가스 배출계수(tCO2eq/MWh)"
# 단위 동치: 1 tCO2eq/MWh = 1 kgCO2eq/kWh (계산 시 변환 불필요)
# 적용 방식: 사용자의 실제 난방 에너지 수요(kWh) 또는 HP 전력 사용량(kWh)에 곱해 kgCO₂ 산출.
# 이전 버전(만원당 kg 환산)보다 사용자 지역/기후존/COP가 자동 반영되어 정확합니다.
# 실제 값은 load_tariff_xlsx()에서 동적 로드됨.

# ── Sheet2 (kWh 기반 물리 데이터) 행 매핑 ──
# 표준 가구(중부2, 거창군 32평, 연 65만원) 기준의 월별 에너지 흐름.
#
# [신규 엑셀 구조 — 2단계 모델]
# 행 43: 실제 난방 에너지 사용량 (kWh) — 사용자가 실제 소비하는 연료량
#        도시가스: (연간요금 - 기본료×12) ÷ 단가 × 난방비중
#        등유/LPG: 연간요금 ÷ 단가
# 행 44: 유효 열 수요 (kWh) — 효율 적용 후 실제 난방 열량 (HP 변환 기준)
#        = 행 43 × 기기 효율
# 행 46-49: 월별 유효 열 수요 (연료별) = 행 44 × 광역시도 월별 비중
# 행 51:    월별 COP (기후존별)
# 행 54-57: 월별 HP 전력 사용량 (연료별) = 행 46-49 ÷ 행 51
SHEET2_FUEL_INPUT_ROW   = 43   # 실제 난방 에너지 사용량 (연료 입력)
SHEET2_HEAT_DEMAND_ROW  = 44   # 유효 열 수요 (HP 변환 기준)
SHEET2_MONTHLY_DEMAND_ROWS = {  # 월별 유효 열 수요 (Excel 행 46-49)
    "도시가스(콘덴싱)": 46, "도시가스(일반)": 47, "등유": 48, "LPG": 49,
}
SHEET2_HP_KWH_ROWS = {          # 월별 HP 전력 사용량 (Excel 행 54-57)
    "도시가스(콘덴싱)": 54, "도시가스(일반)": 55, "등유": 56, "LPG": 57,
}
SHEET2_CO2_DATA_START_ROW = 66  # 15년 CO₂ 데이터 시작 (행 66 = 2026년)

# ── HP 배출계수의 분자 (한국 전력 그리드 평균 배출계수) ──
# 엑셀 Sheet2!I10 = 0.4173 / sCOP[zone]   (2025년)
# 엑셀 Sheet2!J10 = (83.1/624.5) / sCOP[zone]  (2038년, 그리드 청정화 반영)
# 사용자 기후존(중부1/중부2/남부/제주)에 따라 sCOP가 다르므로 HP 배출계수도 동적으로 계산.
GRID_EF_2025_KGKWH = 0.4173          # kg CO2 / kWh — 한국 전력 그리드 (2025년)
GRID_EF_2038_KGKWH = 83.1 / 624.5    # ≈ 0.13306 — 2038년 그리드 (청정화 반영)

# ── Sheet3 광역 시도별 월별 난방 비중 매핑 ──
# 사이트의 약식 시도명 → Sheet3의 정식 명칭
REGION_NAME_MAP = {
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시",
    "인천": "인천광역시", "광주": "광주광역시", "대전": "대전광역시",
    "울산": "울산광역시", "세종": "세종특별자치시",
    "경기도":   "경기도",       "강원도":   "강원특별자치도",
    "충청북도": "충청북도",     "충청남도": "충청남도",
    "전라북도": "전북특별자치도", "전라남도": "전라남도",
    "경상북도": "경상북도",     "경상남도": "경상남도",
    "제주도":   "제주특별자치도",
}

# ── 기초지자체별 zone 오버라이드 (Sheet2 B49 수식 로직) ──
# 광역 시도만으로는 정확하지 않은 지역 — 기초지자체까지 봐야 함
COASTAL_GANGWON   = {"고성군", "속초시", "양양군", "강릉시", "동해시", "삼척시"}  # 강원 해안 → 중부2
NORTHERN_GYEONGGI = {"연천군", "포천시", "가평군", "남양주시", "의정부시", "양주시", "동두천시", "파주시"}  # 경기 북부 → 중부1
SOUTHERN_GYEONGBUK = {"울진군", "영덕군", "포항시", "경주시", "청도군", "경산시"}  # 경북 남쪽 → 남부
NORTHERN_GYEONGBUK = {"봉화군", "청송군"}  # 경북 북쪽 → 중부1
INNER_GYEONGNAM = {"거창군", "함양군"}  # 경남 내륙 → 중부2 (다른 경남은 남부)

# ── 보조금 (만원) ── 난방 전기화 사업: 설치비 1,000만원의 70% 지원
SUBSIDY_TOTAL = 700

# ── 슈퍼유저요금 상수 (엑셀 전기요금!B11·B20 명세) ──
# 동계(12·1·2월)·하계(7·8월) 월 1,000kWh 초과분에 736.2원/kWh 적용.
# ⚠️ 앱 결과는 엑셀 블록이 아니라 이 Python 청구 함수에서 산출되므로,
#    슈퍼유저요금이 HP 비용에 실제로 반영되려면 반드시 여기 코드에 있어야 한다.
#    (엑셀 블록은 결과 계산에 쓰이지 않으므로 이중계산 위험 없음)
SUPER_USER_THRESHOLD_KWH = 1000     # 초과 기준 사용량
SUPER_USER_RATE          = 736.2    # 초과분 단가 (원/kWh)
SUPER_USER_MONTHS        = (12, 1, 2, 7, 8)  # 동계 12·1·2 / 하계 7·8


# ══════════════════════════════════════════════════════════════════════
# 3. 데이터 로더
# ══════════════════════════════════════════════════════════════════════

@st.cache_data
def load_tariff_xlsx():
    """전기요금완료본.xlsx에서 모든 기준 데이터를 로드.

    [전기요금 시트 — 20개 블록]
    각 블록 = (요금제, 태양광유무, 난방유형) 조합.
    블록 헤더 행: col 19=HP 연합계(원), col 20=기존난방비 연합계(원), col 21=Saving 비율
    헤더+1 ~ 헤더+12 행: 1~12월별 청구 내역 (col 9~18)

    [COP_계산기 시트 — 4개 기후존]
    각 존 3행씩: 월평균기온 / 월 HDD(Tbase=18°C) / 월 COP(난방시간 가중)
    sCOP(HDD가중)는 각 존의 첫 행 col 16

    [Sheet2 — 표준 가구의 물리적 에너지 흐름 (kWh)]
    행 43:    실제 난방 에너지 사용량 (연료 입력, 효율 적용 전)
    행 44:    유효 열 수요 (효율 적용 후, HP 변환 기준)
    행 46-49: 월별 유효 열 수요 (연료별)
    행 51:    월별 COP (기후존별)
    행 54-57: 월별 HP 전력 사용량 (연료별) = 행 46-49 ÷ 행 51
    행 66-80: 15년치 연간 CO₂ 배출량 (tCO2eq)

    Returns:
        (data, error_msg) — 성공 시 (dict, None), 실패 시 (None, str)
    """
    fname = "전기요금완료본.xlsx"
    candidates = [
        fname,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), fname),
        os.path.join(os.getcwd(), fname),
    ]
    fp = next((p for p in candidates if os.path.exists(p)), None)
    if fp is None:
        return None, f"파일을 찾을 수 없습니다: {fname}"

    try:
        wb = load_workbook(fp, data_only=True)

        # ── 전기요금 시트 — 20개 블록 ──
        ws = wb["전기요금"]
        blocks = {}
        for key, hr in EXCEL_BLOCK_HEADERS.items():
            blocks[key] = {
                "hp_annual_won":       float(ws.cell(row=hr, column=19).value or 0),
                "existing_annual_won": float(ws.cell(row=hr, column=20).value or 0),
                "saving_ratio":        float(ws.cell(row=hr, column=21).value or 0),
                "monthly": [
                    {
                        "청구요금합계": ws.cell(row=hr+1+m, column=17).value or 0,
                        "HP전기요금":   ws.cell(row=hr+1+m, column=18).value or 0,
                    }
                    for m in range(12)
                ],
            }

        # ── COP_계산기 시트 — 기후존별 sCOP / HDD / 월별 COP / 월평균기온 ──
        ws_cop = wb["COP_계산기"]
        zone_rows = {"중부1": 12, "중부2": 16, "남부": 20, "제주": 24}
        scop, hdd, monthly_cop, monthly_temp = {}, {}, {}, {}
        for zone, r_temp in zone_rows.items():
            r_hdd, r_cop = r_temp + 1, r_temp + 2
            scop[zone] = float(ws_cop.cell(row=r_temp, column=16).value or 0)  # raw — CO2 계산용 정확값
            monthly_temp[zone] = [float(ws_cop.cell(row=r_temp, column=3+m).value or 0) for m in range(12)]
            hdd[zone]          = [float(ws_cop.cell(row=r_hdd,  column=3+m).value or 0) for m in range(12)]
            monthly_cop[zone]  = [float(ws_cop.cell(row=r_cop,  column=3+m).value or 0) for m in range(12)]

        # ── Sheet2 — 연료별 월별 유효 열 수요 / HP 전력 사용량 (kWh) ──
        # SHEET2_MONTHLY_DEMAND_ROWS: 행 46-49 (월별 유효 열 수요)
        # SHEET2_HP_KWH_ROWS:         행 54-57 (월별 HP 전력 사용량)
        ws_s2 = wb["Sheet2"]
        kwh_demand = {fuel: [float(ws_s2.cell(row=r, column=3+m).value or 0) for m in range(12)]
                      for fuel, r in SHEET2_MONTHLY_DEMAND_ROWS.items()}
        kwh_hp     = {fuel: [float(ws_s2.cell(row=r, column=3+m).value or 0) for m in range(12)]
                      for fuel, r in SHEET2_HP_KWH_ROWS.items()}

        # ── Sheet2 핵심 파라미터 (행 43 수식 복제용) ──
        # ⚠️ 행 9·10은 col 5-8 레이아웃 (헤더 행 8: col5=콘덴싱, col6=일반, col7=등유, col8=LPG, col9-10=HP)
        # ⚠️ 행 40·42·14는 col 3-6 레이아웃 (헤더 행 39: col3=콘덴싱, col4=일반, col5=등유, col6=LPG)
        # 같은 시트지만 두 종류의 레이아웃이 섞여있으므로 행마다 시작 col 다름!
        fuel_order = ["도시가스(콘덴싱)", "도시가스(일반)", "등유", "LPG"]
        sheet2_params = {
            "efficiency":    {f: float(ws_s2.cell(row=9,  column=5+i).value or 0) for i, f in enumerate(fuel_order)},
            "base_fee":      {f: float(ws_s2.cell(row=40, column=3+i).value or 0) for i, f in enumerate(fuel_order)},
            "rate":          {f: float(ws_s2.cell(row=42, column=3+i).value or 0) for i, f in enumerate(fuel_order)},
            "heating_share": float(ws_s2.cell(row=14, column=3).value or 0.85),
        }

        # ── Sheet2 행 10 — 기기 온실가스 배출계수 (tCO2eq/MWh = kgCO2eq/kWh) ──
        # 적용 대상이 연료/HP에 따라 다름:
        #   - 연료 (콘덴싱/일반/등유/LPG): 행 43 (실제 난방 에너지 사용량 kWh) × 배출계수
        #     예) 콘덴싱: 6068 kWh × 0.2185 = 1326 kg = 1.33 t (행 66 col 3과 일치)
        #   - 히트펌프: 행 44 (유효 열 수요 kWh) × HP 배출계수
        #     예) 5582 kWh × 0.1312 = 732 kg = 0.73 t (행 66 col 7과 일치)
        # HP 배출계수는 2025년/2038년 두 시점 제공 (그리드 청정화 반영).
        # ⚠️ 행 9와 동일하게 col 5-8 레이아웃 (행 8 헤더: col5=콘덴싱~col8=LPG, col9-10=HP)
        emission_factors_fuel = {
            "도시가스(콘덴싱)": float(ws_s2.cell(row=10, column=5).value or 0),
            "도시가스(일반)":   float(ws_s2.cell(row=10, column=6).value or 0),
            "등유":            float(ws_s2.cell(row=10, column=7).value or 0),
            "LPG":             float(ws_s2.cell(row=10, column=8).value or 0),
        }
        emission_factor_hp_2025 = float(ws_s2.cell(row=10, column=9).value or 0)
        emission_factor_hp_2038 = float(ws_s2.cell(row=10, column=10).value or 0)

        # ── Sheet2 행 66-80 — 15년 연간 온실가스 배출량 (tCO2eq, 표준가구 기준) ──
        # 엑셀 원본은 18년치(행 66-83, 2026-2043)지만 앱은 앞 15년(2026-2040)만 사용
        # 연료(콘덴싱/일반/등유/LPG)는 매년 동일, HP는 그리드 청정화로 매년 감소
        # 검증: 행 66 col 3 = 1.326 t = 행 43 콘덴싱(6068 kWh) × 0.2185 kg/kWh ÷ 1000
        emission_15yr = {}
        for fuel, col in [("도시가스(콘덴싱)", 3), ("도시가스(일반)", 4),
                           ("등유", 5), ("LPG", 6), ("히트펌프", 7)]:
            emission_15yr[fuel] = [float(ws_s2.cell(row=SHEET2_CO2_DATA_START_ROW+y, column=col).value or 0)
                                    for y in range(15)]

        # ── 표준가구 HP 연간 kWh — 행 54 (콘덴싱→HP 월별) 합계 ──
        # CO2 동적 계산용: 행 66~80(표준 베이크 값)을 사용자 HP kWh에 비례 스케일링
        # Sheet2!I9 (0.131)는 행 66의 실효 배출계수(0.448)와 불일치하므로
        # 베이크 데이터를 그대로 비례 사용하는 것이 가장 안전
        standard_hp_annual_kwh = sum(
            float(ws_s2.cell(row=54, column=3+m).value or 0) for m in range(12)
        )

        # ── Sheet3 — 광역시도별 월별 난방 비중 (17개 시도 + 전국) ──
        ws_s3 = wb["Sheet3"]
        region_ratios = {}
        for r in range(2, 20):
            name = ws_s3.cell(row=r, column=1).value
            if not name: continue
            ratios = [float(ws_s3.cell(row=r, column=3+m).value or 0) for m in range(12)]
            region_ratios[name] = ratios

        # ── 전기요금 시트 — 가전 월별 사용량 (행 23-34, col 3) ──
        # 22년 전국 가구패널 기준 평균 가전 전력 사용량
        appliance_kwh = [float(ws.cell(row=23+m, column=3).value or 0) for m in range(12)]

        # ── Sheet2 행 25 — 월별 전기요금 비중 (사용자 입력 1월 kWh → 12개월 분배용) ──
        elec_monthly_ratios = [float(ws_s2.cell(row=25, column=3+m).value or 0) for m in range(12)]

        # ── 전기요금 시트 — 17개 시도별 월별 태양광 발전량 (col 37=시도, col 39-50=1~12월) ──
        # 1kW 패널당 발전량 kWh
        solar_kwh_by_region = {}
        for r in range(5, 22):
            name = ws.cell(row=r, column=37).value
            if not name: continue
            solar_kwh_by_region[name] = [float(ws.cell(row=r, column=39+m).value or 0) for m in range(12)]

        return {
            "blocks":         blocks,
            "scop":           scop,
            "hdd":            hdd,
            "monthly_cop":    monthly_cop,
            "monthly_temp":   monthly_temp,
            "kwh_demand":     kwh_demand,
            "kwh_hp":         kwh_hp,
            "sheet2_params":  sheet2_params,
            "region_ratios":  region_ratios,
            "appliance_kwh":  appliance_kwh,
            "elec_monthly_ratios": elec_monthly_ratios,
            "solar_kwh":      solar_kwh_by_region,
            "emission_factors_fuel":   emission_factors_fuel,
            "emission_factor_hp_2025": emission_factor_hp_2025,
            "emission_factor_hp_2038": emission_factor_hp_2038,
            "emission_15yr":  emission_15yr,
            "standard_hp_annual_kwh": standard_hp_annual_kwh,
        }, None
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════════════
# 4. 계산 함수
# ══════════════════════════════════════════════════════════════════════

def map_region_to_zone(region, sub_region=""):
    """광역+기초지자체 조합으로 기후존 결정 (엑셀 Sheet2 B49 수식 로직 복제).

    더 정밀한 분류: 강원 해안(따뜻함) / 경기 북부(추움) / 경북 남쪽(따뜻함) 등 반영.
    """
    # 제주
    if region == "제주도":
        return "제주"
    # 남부 (해안 따뜻한 광역)
    if region in {"부산", "대구", "울산", "광주", "전라남도"}:
        return "남부"
    # 남부 (경북 남쪽)
    if region == "경상북도" and sub_region in SOUTHERN_GYEONGBUK:
        return "남부"
    # 남부 (경남, 단 거창·함양 제외)
    if region == "경상남도" and sub_region not in INNER_GYEONGNAM:
        return "남부"
    # 중부1 (강원 산악)
    if region == "강원도" and sub_region not in COASTAL_GANGWON:
        return "중부1"
    # 중부1 (경기 북부)
    if region == "경기도" and sub_region in NORTHERN_GYEONGGI:
        return "중부1"
    # 중부1 (충북 제천)
    if region == "충청북도" and sub_region == "제천시":
        return "중부1"
    # 중부1 (경북 북쪽 봉화·청송)
    if region == "경상북도" and sub_region in NORTHERN_GYEONGBUK:
        return "중부1"
    # 나머지 = 중부2 (서울/대전/세종/인천/충남/전북/강원해안/경기 외/충북 외/경남 거창함양/경북 외)
    return "중부2"


def get_block_key(tariff_label, heating_ui):
    """UI 라벨로부터 엑셀 블록 키 결정.

    Returns: (block_key, tariff, solar_flag)
    """
    tariff, solar = TARIFF_LABEL_MAP[tariff_label]
    heating = HEATING_TYPE_MAP[heating_ui]
    return (tariff, solar, heating), tariff, solar



def get_hp_specs(h_size_pyung):
    """전용면적(평) → (설치 공간 비유, 치수, HP 용량).

    적정 용량 기준:
      • 14평 이하   → 8 kW  (냉장고 크기)
      • 15~25평     → 12 kW (워시타워 1대 크기)
      • 26~35평     → 16 kW (보일러실 크기)
      • 36~40평     → 20 kW (대형 보일러실 크기)
      • 40평 초과   → 상담 필요 (단일 기기로 부족 — 다중 설치/맞춤 설계 권장)
    """
    if h_size_pyung <= 14:
        return ("냉장고 크기",       "700 × 1,800 mm",   "8 kW")
    if h_size_pyung <= 25:
        return ("워시타워 1대 크기", "800 × 1,115 mm",   "12 kW")
    if h_size_pyung <= 35:
        return ("보일러실 크기",     "1,120 × 1,666 mm", "16 kW")
    if h_size_pyung <= 40:
        return ("대형 보일러실 크기", "1,380 × 1,700 mm", "20 kW")
    return     ("상담 필요",         "—",                "상담 필요")


def get_hp_capacity_kw(h_size_pyung):
    """전용면적(평) → HP 용량 숫자(kW). 계산용.

    14평 이하 8 / ~25평 12 / ~35평 16 / ~40평 20 / 40평 초과 24(상담 가정).
    40평 초과는 단일 기기 한계를 넘어 '상담 필요'로 표시하지만,
    계약전력·요금 계산이 멈추지 않도록 내부적으로는 24kW를 가정한다.
    """
    if h_size_pyung <= 14:   return 8
    if h_size_pyung <= 25:   return 12
    if h_size_pyung <= 35:   return 16
    if h_size_pyung <= 40:   return 20
    return 24


def calc_monthly_stats(monthly_ex_man, monthly_hp_man, monthly_ratios,
                       annual_ex_kg, annual_hp_kg):
    """월별 절감액·누적·절감률·CO₂ 한 번에 계산.

    CO₂는 엑셀 Sheet2 행 65 (연간 온실가스 배출량 tCO2eq)을 그대로 사용하고,
    광역시도 월별 난방 비중(monthly_ratios)으로 안분.

    Args:
        monthly_ex_man, monthly_hp_man:  월별 비용 (만원)
        monthly_ratios:                  광역시도 월별 비중 (합=1, Sheet3)
        annual_ex_kg:                    연간 기존 보일러 배출량 (kg, 엑셀 표준 가구)
        annual_hp_kg:                    연간 HP 배출량 (kg, 엑셀 표준 가구)
    """
    annual_saving_kg = max(0, annual_ex_kg - annual_hp_kg)
    savings, cumulative, savings_pct, co2 = [], [], [], []
    cum = 0.0
    for ex, hp, ratio in zip(monthly_ex_man, monthly_hp_man, monthly_ratios):
        sav = round(ex - hp, 2)
        cum += sav
        savings.append(sav)
        cumulative.append(round(cum, 2))
        # 비난방월(기존 난방비 0)은 비율 의미 없음 → "-"
        savings_pct.append(f"{round(sav/ex*100, 1)}%" if ex > 0 else "-")
        # 월별 CO₂ 절감 = 연간 절감 × 그 달 비중
        co2.append(round(annual_saving_kg * ratio, 1))
    return {
        "savings":    savings,
        "cumulative": cumulative,
        "pct":        savings_pct,
        "co2":        co2,
    }


def calc_annual_co2_emissions(user_annual_cost_won, user_heat_demand_kwh,
                              user_heating_share, fuel_key,
                              sheet2_params, emission_factors_fuel,
                              emission_factor_hp_2025, emission_factor_hp_2038,
                              year_idx=0):
    """연간 CO₂ 배출량 — 엑셀 Sheet2 G66~G80 공식 정확 복제.

    엑셀 공식 (Sheet2):
      • C66/D66/E66/F66 (1년차, 연료별) = 행43/1000 × 배출계수 (행10)
      • G66 (1년차 HP, 2026) = user_heat_demand_kwh / 1000 × I10  (단위: tCO2)
      • G78 (13년차 HP, 2038) = user_heat_demand_kwh / 1000 × J10
      • G67~G77 (2~12년차): G66 + (idx-1) × (G78 - G66) / 12   ← 선형 보간
      • G79 (14년차): G78 - G78/12
      • G80 (15년차): G79 - G78/12  =  G78 - 2 × G78/12

    중요: HP CO2는 **유효 열 수요(kWh)** 기반 계산 (HP가 만들어야 할 열량).
         **HP 전력 사용량**과는 다르다 (전력 = 열 ÷ COP).
         이전 버전은 후자를 썼기 때문에 약 1/COP(≈1/3)로 과소 추정됨.

    Args:
        user_annual_cost_won:    사용자 연간 난방비
        user_heat_demand_kwh:    유효 열 수요 (kWh) — 엑셀 C44/D44/E44/F44에 해당
        user_heating_share:      취사기기 보정 — 도시가스/LPG 0.8475, 인덕션 1.0
        fuel_key:                현재 사용 연료
        sheet2_params:           base_fee, rate
        emission_factors_fuel:   연료별 배출계수 (Sheet2 행 10 col 5-8, kg/kWh)
        emission_factor_hp_2025: I10 ≈ 0.131 kg/kWh
        emission_factor_hp_2038: J10 ≈ 0.0418 kg/kWh
        year_idx:                0=2026, 1~14=2027~2040

    참고: 엑셀 패턴에는 태양광 자가발전 차감이 없음 (그리드 청정화는 매년 자연 감소).
         태양광 효과는 전력비 절감 측면에서만 반영되고, CO2는 엑셀과 동일하게 처리.
    """
    base_fee = sheet2_params["base_fee"]
    rate     = sheet2_params["rate"]

    # 5개 연료별 — "같은 연간 난방비를 각 연료에 썼을 때" CO2 (kg)
    by_fuel = {}
    for fuel in ["도시가스(콘덴싱)", "도시가스(일반)", "등유", "LPG"]:
        if rate.get(fuel, 0) <= 0:
            by_fuel[fuel] = 0
            continue
        fuel_kwh = (user_annual_cost_won - base_fee[fuel] * 12) / rate[fuel]
        if "도시가스" in fuel:
            fuel_kwh *= user_heating_share
        fuel_kwh = max(0, fuel_kwh)
        by_fuel[fuel] = fuel_kwh * emission_factors_fuel.get(fuel, 0)  # kg

    # HP CO2 — 엑셀 G66~G80 패턴 (단위: kg)
    co2_y1  = user_heat_demand_kwh * emission_factor_hp_2025   # G66 (× 1000)
    co2_y13 = user_heat_demand_kwh * emission_factor_hp_2038   # G78 (× 1000)

    if year_idx <= 12:        # G66~G78 (1~13년차 = 2026~2038) 선형 보간
        hp_kg = co2_y1 + year_idx * (co2_y13 - co2_y1) / 12
    elif year_idx == 13:      # G79 (14년차 = 2039)
        hp_kg = co2_y13 - co2_y13 / 12
    else:                      # G80 (15년차 = 2040)
        hp_kg = co2_y13 - 2 * co2_y13 / 12

    by_fuel["히트펌프"] = max(0, hp_kg)

    ex_kg = by_fuel.get(fuel_key, 0)
    hp_kg = by_fuel["히트펌프"]
    return {
        "ex_kg":     ex_kg,
        "hp_kg":     hp_kg,
        "saving_kg": max(0, ex_kg - hp_kg),
        "by_fuel":   by_fuel,
    }


def calc_fuel_input_kwh(annual_cost_won, fuel_key, sheet2_params, heating_share=None):
    """Sheet2 행 43 수식 — 실제 난방 에너지 사용량 (kWh).

    사용자가 1년 동안 실제로 소비하는 연료 에너지량 (효율 적용 전).
    가스 고지서나 등유 구매량으로 환산 가능한 값.

    수식 (도시가스):  (연간요금 - 기본요금×12) ÷ 단가 × 난방비중
    수식 (등유/LPG):  연간요금 ÷ 단가  (기본료 0, 난방비중 미적용)

    Args:
        heating_share: 엑셀 Sheet2!C14 — 도시가스 사용 중 난방이 차지하는 비중
                       None이면 sheet2_params의 기본값 사용
                       엑셀 공식: 취사가 도시가스/LPG면 0.8475, 인덕션이면 1.0
    """
    base_fee = sheet2_params["base_fee"][fuel_key]
    rate     = sheet2_params["rate"][fuel_key]
    share    = heating_share if heating_share is not None else sheet2_params["heating_share"]
    if rate <= 0: return 0
    fuel_input_kwh = (annual_cost_won - base_fee * 12) / rate
    if "도시가스" in fuel_key:                # 도시가스만 난방비중 적용
        fuel_input_kwh *= share
    return max(0, fuel_input_kwh)


def calc_heat_demand_kwh(fuel_input_kwh, fuel_key, sheet2_params):
    """Sheet2 행 44 수식 — 유효 열 수요 (kWh).

    실제 난방으로 쓰이는 열량 (효율 적용 후) — HP가 만들어야 할 열량과 동일.
    수식: 실제 사용량(행 43) × 기기 효율
    """
    eff = sheet2_params["efficiency"][fuel_key]
    return fuel_input_kwh * eff


def calc_kwh_data(fuel_input_annual, heat_demand_annual, monthly_ratios, monthly_cop_zone):
    """Sheet2 행 46-49, 54-57 수식 복제 — 월별 kWh 계산.

    Args:
        fuel_input_annual:   행 43 결과 — 실제 연료 사용량 (kWh, 표시용)
        heat_demand_annual:  행 44 결과 — 유효 열 수요 (kWh, HP 변환 기준)
        monthly_ratios:      사용자 광역시도의 월별 난방 비중 (Sheet3, 합=1)
        monthly_cop_zone:    사용자 기후존의 월별 COP

    Returns: dict
        fuel_input_annual: 행 43 — 기존 보일러 연간 연료 사용량
        heat_demand_annual: 행 44 — 유효 열 수요 (= HP가 만들어야 할 열량)
        monthly_demand:    행 46-49 — 월별 유효 열 수요
        monthly_hp:        행 54-57 — 월별 HP 전력 사용량
        annual_hp:         연간 HP 전력 사용량
        efficiency:        실 에너지 비교 배수 (= 연료 사용량 ÷ HP 전력 사용량)
    """
    monthly_demand = [round(heat_demand_annual * r, 1) for r in monthly_ratios]
    monthly_hp = [round(d / cop, 1) if cop > 0 else 0.0
                  for d, cop in zip(monthly_demand, monthly_cop_zone)]
    annual_hp = round(sum(monthly_hp), 0)
    # 사용자 관점 효율: "기존엔 X kWh 쓰던 걸 HP로는 Y kWh로" — 실제 에너지 비교
    efficiency = round(fuel_input_annual / annual_hp, 1) if annual_hp > 0 else 0.0
    return {
        "fuel_input_annual":  round(fuel_input_annual, 0),    # 행 43 (보일러 연료)
        "heat_demand_annual": round(heat_demand_annual, 0),   # 행 44 (유효 열 수요)
        "monthly_demand":     monthly_demand,                 # 행 46-49
        "monthly_hp":         monthly_hp,                     # 행 54-57
        "annual_hp":          annual_hp,
        "efficiency":         efficiency,
    }


def calc_progressive_billing(usage_kwh, month, solar_offset=0):
    """누진제 적용 월별 청구액 계산 (전기요금 시트 J~Q열 수식 그대로 복제).

    수식 출처: 전기요금!J6:Q6 (1월 누진제 태X 기준).

    슈퍼유저요금 반영 (엑셀 전기요금!B11·B20 명세):
      동계(12·1·2월)·하계(7·8월)에 한해 월 1,000kWh 초과분은 736.2원/kWh 적용.
      히트펌프 난방으로 겨울 사용량이 1,000kWh를 넘는 경우를 정확히 반영한다.
      (앱 결과는 이 함수에서 산출되므로 슈퍼유저요금은 여기서 가산해야 한다.)

    Args:
        usage_kwh:    월별 총 사용량 (HP + 가전)
        month:        1~12
        solar_offset: 태양광 자가발전 차감량 (kWh, 태양광 설치 시)

    Returns: 청구요금합계 (원, 10원 단위 절사)
    """
    # 태양광 자가발전 차감 (누진제 태O 케이스)
    actual = max(usage_kwh - solar_offset, 0)

    # 누진 단계 (7-8월 하계는 단계 다름)
    if month in (7, 8):
        step1_max, step2_max = 300, 450
        step1_acc, step2_acc = 36000, 68190   # 누적 단가
    else:
        step1_max, step2_max = 200, 400
        step1_acc, step2_acc = 24000, 66920

    # 기본요금
    if actual <= step1_max:    base_fee = 910
    elif actual <= step2_max:  base_fee = 1600
    else:                       base_fee = 7300

    # 슈퍼유저요금 발동 여부 — 동·하계 & 1,000kWh 초과
    super_on = month in SUPER_USER_MONTHS and actual > SUPER_USER_THRESHOLD_KWH

    # 사용량요금 (단계별 누진)
    if actual <= step1_max:
        usage_fee = int(actual * 120)
    elif actual <= step2_max:
        usage_fee = int(step1_acc + (actual - step1_max) * 214.6)
    elif not super_on:
        usage_fee = int(step2_acc + (actual - step2_max) * 307.3)
    else:
        # 1,000kWh까지는 최고누진(307.3원) 누적, 초과분만 736.2원/kWh
        acc_1000  = step2_acc + (SUPER_USER_THRESHOLD_KWH - step2_max) * 307.3
        usage_fee = int(acc_1000 + (actual - SUPER_USER_THRESHOLD_KWH) * SUPER_USER_RATE)

    climate_fee = int(actual * 9)        # 기후환경요금
    fuel_adj    = int(actual * 5)        # 연료비조정요금
    total_fee   = base_fee + usage_fee + climate_fee + fuel_adj  # 전기요금계
    vat         = round(total_fee * 0.1)                          # 부가가치세 10%
    fund        = (int(total_fee * 0.027) // 10) * 10             # 기금 2.7% (10원 절사)
    billing     = ((total_fee + vat + fund) // 10) * 10            # 청구합계 (10원 절사)
    return billing


def calc_hp_billing_progressive(monthly_hp_kwh, monthly_appliance_kwh, monthly_solar_kwh=None):
    """누진제 케이스에서 HP 분리 청구액 12개월 계산 — 증분비용 방식.

    [2026-06 변경] 기존 사용량 비율 분배(Q × HP/(HP+가전))에서
    증분비용 방식으로 교체:
        HP 청구액 = 청구액(HP+가전) − 청구액(가전만)
    가전만으로는 1,000kWh를 넘지 않으므로, HP가 밀어올린 고누진·슈퍼유저
    구간 비용이 온전히 HP에 귀속된다. (비율분배는 이 비싼 구간을 희석해 과소계상)

    Args:
        monthly_hp_kwh:        12개월 HP 전력 사용량
        monthly_appliance_kwh: 12개월 가전 평균 사용량
        monthly_solar_kwh:     12개월 태양광 발전량 (있으면 차감, 누진제 태O)

    Returns: 12개월 HP 청구액 (원) 리스트
    """
    hp_won_monthly = []
    for m in range(12):
        hp_kwh = monthly_hp_kwh[m]
        gadget_kwh = monthly_appliance_kwh[m]
        solar = monthly_solar_kwh[m] if monthly_solar_kwh else 0

        # 증분 = (HP+가전) 청구액 − (가전만) 청구액
        bill_with = calc_progressive_billing(hp_kwh + gadget_kwh, m+1, solar_offset=solar)
        bill_base = calc_progressive_billing(gadget_kwh,          m+1, solar_offset=solar)
        hp_won_monthly.append(max(0, bill_with - bill_base))

    return hp_won_monthly


# ─── 일반용 (HP 전용 미터) 단가 ───
GENERAL_RATE_BY_MONTH = {1:119, 2:119, 11:119, 12:119,        # 겨울
                         3:91.9, 4:91.9, 5:91.9, 9:91.9, 10:91.9,  # 봄가을
                         6:132.4, 7:132.4, 8:132.4}            # 여름
GENERAL_BASE_PER_KW = 6160      # 기본료 (원/kW)
CONTRACT_COP_DIVISOR = 3        # 엑셀 D110: 계약전력 산정용 COP

def calc_general_billing(hp_kwh, month, contract_kw):
    """일반용 (HP 전용 미터) 청구액 — 전기요금!J110~Q110 수식 복제.

    HP만 별도 미터로 측정. R = Q (HP 분리 안 함).
    """
    base_fee = contract_kw * GENERAL_BASE_PER_KW
    rate = GENERAL_RATE_BY_MONTH[month]
    usage_fee = int(hp_kwh * rate)              # ROUNDDOWN
    climate_fee = int(hp_kwh * 9)
    fuel_adj    = int(hp_kwh * 5)
    total_fee   = base_fee + usage_fee + climate_fee + fuel_adj
    vat         = round(total_fee * 0.1)
    fund        = (int(total_fee * 0.027) // 10) * 10
    billing     = ((total_fee + vat + fund) // 10) * 10
    return billing


# ─── 계시별 단가표 ───
TOU_BASE_PER_KW = 4310       # 기본료 (원/kW)
TOU_GADGET_KW   = 3          # 가전 추가 계약용량 (kW)
TOU_RATES = {
    "spring":        (125.8, 153.8, 172.4),  # 3-5, 9-10월 (경부하/중간/최대)
    "winter_summer": (138.7, 184.7, 220.5),  # 1-2, 6-8, 11-12월
}
GADGET_TOU = (0.30, 0.30, 0.40)   # 가전 시간대 분포 (모든 월 동일)

# HP의 시간대 분포 — 월별로 다름 (1월~12월)
HP_TOU = [
    (0.4337, 0.2876, 0.2787), (0.4403, 0.2880, 0.2716), (0.4416, 0.2826, 0.2758),
    (0.4251, 0.2779, 0.2970), (0.3844, 0.3044, 0.3112), (0.3192, 0.3886, 0.2922),
    (0.2680, 0.4376, 0.2944), (0.2688, 0.4515, 0.2797), (0.3032, 0.4048, 0.2920),
    (0.4234, 0.2810, 0.2956), (0.4443, 0.2722, 0.2835), (0.4382, 0.2818, 0.2800),
]
# 태양광의 시간대 분배 — 월별로 다름 (계시별 태O용)
SOLAR_TOU = [
    (0.000062, 0.954981, 0.044957), (0.001791, 0.921596, 0.076613),
    (0.011490, 0.896370, 0.092139), (0.026448, 0.868176, 0.105375),
    (0.041831, 0.844735, 0.113434), (0.047395, 0.816742, 0.135862),
    (0.036986, 0.818888, 0.144126), (0.032913, 0.858900, 0.108187),
    (0.024481, 0.895608, 0.079911), (0.013153, 0.939251, 0.047596),
    (0.003291, 0.971714, 0.024994), (0.000353, 0.977840, 0.021807),
]

def calc_tou_billing(hp_kwh, gadget_kwh, month, contract_kw, solar_kwh=0):
    """계시별 청구액 — 전기요금!J162~Q162 (태X) / J214~Q214 (태O) 수식 복제."""
    season = "spring" if month in (3,4,5,9,10) else "winter_summer"
    rates = TOU_RATES[season]
    hp_r = HP_TOU[month-1]
    solar_r = SOLAR_TOU[month-1] if solar_kwh > 0 else (0, 0, 0)

    base_fee = TOU_BASE_PER_KW * (contract_kw + TOU_GADGET_KW)

    usage_fee = sum(
        max(hp_kwh * hp_r[i] + gadget_kwh * GADGET_TOU[i] - solar_kwh * solar_r[i], 0) * rates[i]
        for i in range(3)
    )
    usage_fee = int(usage_fee)   # ROUNDDOWN

    total_after_solar = max(hp_kwh + gadget_kwh - solar_kwh, 0)
    climate_fee = int(total_after_solar * 9)
    fuel_adj    = int(total_after_solar * 5)

    total_fee = base_fee + usage_fee + climate_fee + fuel_adj
    vat       = round(total_fee * 0.1)
    fund      = (int(total_fee * 0.027) // 10) * 10
    billing   = ((total_fee + vat + fund) // 10) * 10
    return billing


def calc_dynamic_result(tariff_label, monthly_hp_kwh, monthly_appliance_kwh,
                        monthly_solar_kwh, hp_capacity_kw, ex_annual_won):
    """동적 계산 - 사용자 지역 기준 5개 요금제 모두 처리.

    엑셀 전기요금 시트의 R/S/T/U 수식을 Python으로 복제.
    HP 월별 kWh + 가전 + 태양광을 받아 요금제별 월별 청구액 산정.

    [2026-06 변경] 누진제 HP 분리를 증분비용 방식으로 교체:
        HP 청구액 = 청구액(HP+가전) − 청구액(가전만)
    슈퍼유저 구간(calc_progressive_billing 내부 반영)이 HP에 온전히 귀속된다.

    Args:
        tariff_label:           "누진제 (태양광 미설치)" 등 5개 라벨
        monthly_hp_kwh:         12개월 HP 전력 사용량 (사용자 지역 기준)
        monthly_appliance_kwh:  12개월 가전 평균 사용량
        monthly_solar_kwh:      12개월 태양광 발전량 (사용자 광역시도 기준)
        hp_capacity_kw:         HP 용량 (get_hp_specs 결과)
        ex_annual_won:          사용자 추정 연간 난방비 (원)

    Returns: dict with monthly_won, monthly_man, hp_annual_man, ex_annual_man, saving_man, saving_ratio
    """
    import math

    # 요금제·태양광 분기
    tariff_kind, solar_on = TARIFF_LABEL_MAP[tariff_label]
    contract_kw = math.ceil(hp_capacity_kw / CONTRACT_COP_DIVISOR)

    monthly_hp_won = []
    for m in range(12):
        hp = monthly_hp_kwh[m]
        gad = monthly_appliance_kwh[m]
        solar = monthly_solar_kwh[m] if (solar_on == "태O" and monthly_solar_kwh) else 0
        total = hp + gad

        if tariff_kind == "누진제":
            # 증분비용 방식: HP가 일으킨 한계 청구액만 귀속
            # = 청구액(HP+가전) − 청구액(가전만). 슈퍼유저 구간이 HP에 온전히 반영됨.
            bill_with = calc_progressive_billing(total, m+1, solar_offset=solar)
            bill_base = calc_progressive_billing(gad,   m+1, solar_offset=solar)
            hp_won = max(0, bill_with - bill_base)

        elif tariff_kind == "일반용":
            # HP 전용 미터 — 가전·태양광 무관
            hp_won = calc_general_billing(hp, m+1, contract_kw)

        elif tariff_kind == "계시별":
            billing = calc_tou_billing(hp, gad, m+1, contract_kw, solar_kwh=solar)
            hp_won = round(billing * hp / total) if total > 0 else 0
        else:
            hp_won = 0

        monthly_hp_won.append(hp_won)

    monthly_man   = [round(w / 10000, 2) for w in monthly_hp_won]
    hp_annual_man = round(sum(monthly_hp_won) / 10000, 1)
    ex_annual_man = round(ex_annual_won / 10000, 1)
    saving_man    = round(ex_annual_man - hp_annual_man, 1)
    saving_ratio  = saving_man / ex_annual_man if ex_annual_man > 0 else 0

    return {
        "monthly_won":    monthly_hp_won,
        "monthly_man":    monthly_man,
        "hp_annual_man":  hp_annual_man,
        "ex_annual_man":  ex_annual_man,
        "saving_man":     saving_man,
        "saving_ratio":   saving_ratio,
    }


def simulate_15yr(net_capex_man, ann_heat_man, ann_hp_man, fuel_inflation_pct, elec_inflation_pct):
    """15년 누적 비용·순이익 시뮬레이션 (인플레이션 복리 적용).

    1년차는 현재 가격(인플레이션 0), 2년차부터 복리 적용 → 지수 (y-1).
    (이전 버전은 1년차부터 ^y로 한 해 더 붙어 순이익이 과대평가되었음.)

    Returns: (years, gas_cum, hp_cum, net_profit, payback_year)
    """
    years = list(range(1, 16))
    gas_cum, hp_cum, net_profit = [], [], []
    gas_total, hp_total = 0.0, float(net_capex_man)
    payback = "15년 초과"

    for y in years:
        gas_total += ann_heat_man * ((1 + fuel_inflation_pct / 100) ** (y - 1))
        hp_total  += ann_hp_man   * ((1 + elec_inflation_pct / 100) ** (y - 1))
        profit = int(gas_total - hp_total)
        gas_cum.append(int(gas_total))
        hp_cum.append(int(hp_total))
        net_profit.append(profit)
        if payback == "15년 초과" and profit > 0:
            payback = f"{y}년차"

    return years, gas_cum, hp_cum, net_profit, payback


# ══════════════════════════════════════════════════════════════════════
# 6. UI — 헤더 및 솔루션 개요
# ══════════════════════════════════════════════════════════════════════

excel_data, load_err = load_tariff_xlsx()
if excel_data:
    tariff_blocks = excel_data["blocks"]
    SCOP_BY_ZONE  = excel_data["scop"]          # 동적 로드된 sCOP
    HDD_MONTHLY   = excel_data["hdd"]           # 동적 로드된 HDD (Tbase=18°C)
    MONTHLY_COP   = excel_data["monthly_cop"]   # 월별 COP (참고용)
    MONTHLY_TEMP  = excel_data["monthly_temp"]  # 월평균 기온 (참고용)
    KWH_DEMAND    = excel_data["kwh_demand"]    # Sheet2: 연료별 월별 난방 수요 (참고용, 거창군 기준)
    KWH_HP        = excel_data["kwh_hp"]        # Sheet2: 연료별 월별 HP 전력 사용량 (참고용)
    SHEET2_PARAMS = excel_data["sheet2_params"] # Sheet2: 연료별 효율/단가/기본요금/난방비중
    REGION_RATIOS = excel_data["region_ratios"] # Sheet3: 광역시도별 월별 난방 비중
    APPLIANCE_KWH = excel_data["appliance_kwh"] # 전기요금 시트: 가전 월별 평균 사용량 (전국, 12개) — 사용자 입력 없을 때 fallback
    ELEC_MONTHLY_RATIOS = excel_data["elec_monthly_ratios"]  # Sheet2 행 25: 월별 전기요금 비중 (합=1)
    SOLAR_KWH     = excel_data["solar_kwh"]     # 전기요금 시트: 17개 시도 월별 1kW당 태양광 발전량
    EMISSION_FACTORS_FUEL   = excel_data["emission_factors_fuel"]    # Sheet2 행 10 col 5-8: 연료별 배출계수 (kg/kWh)
    EMISSION_FACTOR_HP_2025 = excel_data["emission_factor_hp_2025"]  # Sheet2 행 10 col 9: HP 2025년 배출계수
    EMISSION_FACTOR_HP_2038 = excel_data["emission_factor_hp_2038"]  # Sheet2 행 10 col 10: HP 2038년 배출계수 (참고용)
    EMISSION_15YR = excel_data["emission_15yr"] # Sheet2 행 65-79: 표준가구 15년치 연간 배출량
    STANDARD_HP_ANNUAL_KWH = excel_data["standard_hp_annual_kwh"]  # Sheet2 행 54 합계 ≈ 1635 (CO2 스케일링용)

col_title, col_logo = st.columns([6, 1])
with col_title:
    st.title("GESI 히트펌프 경제성·환경성 간이 계산기")
with col_logo:
    if os.path.exists("logo.png"):
        st.image(Image.open("logo.png"), width=120)

# 로드 실패 시 중단
if load_err:
    st.error(f"⚠️ 전기요금완료본.xlsx 로드 실패: {load_err}")
    st.info("repo 루트(또는 app.py 옆)에 `전기요금완료본.xlsx` 파일이 있는지 확인해 주세요.")
    st.stop()

st.markdown("""
<div class='info-box'>
  <h4 class='info-title'>💡 이 계산기가 왜 필요한가요?</h4>

  <div style='margin-bottom:20px;'>
    <p style='color:#0f172a; font-size:1.05rem; font-weight:600; margin-bottom:6px;'>🌍 왜 지금 히트펌프인가요?</p>
    <p class='info-text'>
      우리나라 가정의 난방은 대부분 <b>가스·등유 보일러</b>로 이루어지고 있고, 이는 가정에서 발생하는
      탄소 배출의 가장 큰 원인입니다. 정부는 <b>2050 탄소중립</b> 목표 달성을 위해 친환경 히트펌프
      전환 시 <b>최대 70%의 보조금</b>(일부 지역)을 지원하고 있어요. 가스요금이
      해마다 오르는 만큼, 지금이 우리 집 난방을 바꾸는 게 정말 이득인지 미리 따져볼 좋은 시점입니다.
    </p>
  </div>

  <div style='margin-bottom:20px;'>
    <p style='color:#0f172a; font-size:1.05rem; font-weight:600; margin-bottom:6px;'>🏠 히트펌프(AWHP)가 뭔가요?</p>
    <p class='info-text'>
      <b>공기 중에 있는 열을 모아 난방에 쓰는 친환경 기기</b>입니다. 에어컨이 실내 열을 밖으로 빼내는
      것과 정반대로 작동하는 원리예요. 가스를 태우는 게 아니라 전기로 움직이고, 같은 에너지로
      <b>가스보일러보다 약 3~4배 효율</b>이 나옵니다. 그래서 태양광 설비를 함께 쓰면
      난방비를 거의 0에 가깝게 만들 수도 있습니다.
    </p>
  </div>

  <div>
    <p style='color:#0f172a; font-size:1.05rem; font-weight:600; margin-bottom:6px;'>📝 어떻게 사용하나요?</p>
    <ol style='color:#475569; font-size:1.0rem; line-height:1.85; margin:0; padding-left:22px;'>
      <li>우리 집 <b>지역과 평수</b>를 골라주세요</li>
      <li>지난겨울 <b>1월 난방비와 전기요금</b>, 현재 쓰는 <b>난방 방식</b>을 입력해 주세요
        <span style='color:#94a3b8; font-size:0.9rem;'>(고지서를 참고하시면 가장 정확합니다)</span></li>
      <li>사용 중인 <b>전기 요금제</b>를 5가지 중 하나 골라주세요
        <span style='color:#94a3b8; font-size:0.9rem;'>(태양광 설치 여부 포함)</span></li>
      <li>받을 수 있는 <b>보조금</b>을 체크하세요</li>
      <li><b>'경제성·환경성 분석 실행'</b> 버튼을 누르면 끝!
        절감액·투자 회수 기간이 한 눈에 보입니다.</li>
    </ol>
  </div>

  <div style='margin-top:22px; padding:14px 18px; background:#e6f4f1; border-left:3px solid #0f766e; border-radius:8px;'>
    <p style='color:#134e4a; font-size:0.95rem; line-height:1.6; margin:0;'>
      ℹ️ <b>참고</b>: 본 어플리케이션에서는 히트펌프의 설치 비용(CAPEX)을 <b>1,000만원</b>으로 설정하였습니다.
      <span style='color:#78716c; font-size:0.88rem;'>(국내 기업 평균 견적 기준 — 본체+설치비+부대공사 포함)</span>
    </p>
  </div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
# 7. UI — 입력 섹션
# ══════════════════════════════════════════════════════════════════════

# ── 섹션 1: 대상지 ──
st.markdown('<div class="section-title">1. 대상지 기본 정보</div>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    region = st.selectbox("광역 지자체", list(REGIONS_FULL.keys()), index=0)
with col2:
    sub_region = st.selectbox("기초 지자체", REGIONS_FULL.get(region, ["전체"]))

col3, col4 = st.columns(2)
with col3:
    house_type = st.selectbox("주거 형태",
                              ["단독 주택 / 다가구 주택", "아파트", "연립 / 빌라 / 다세대 주택"])
with col4:
    house_size = st.number_input("전용 면적 (평)", min_value=10, value=30)

zone        = map_region_to_zone(region, sub_region)
dynamic_cop = SCOP_BY_ZONE[zone]   # 로드 시점에 이미 2자리로 round됨 (2.90/3.18/3.43/3.56)

# ── 섹션 2: 에너지 소비 ──
st.markdown('<div class="section-title">2. 에너지 소비 현황</div>', unsafe_allow_html=True)
col_h, col_e = st.columns(2)
with col_h:
    winter_heat_man = st.number_input("동절기(1월) 평균 난방비 (만원)", value=20)
with col_e:
    winter_elec_kwh = st.number_input(
        "동절기(1월) 전기 사용량 (kWh)", value=0, min_value=0,
        help="한전 청구서의 '1월 사용량(kWh)'을 입력해 주세요. 4인 가구 평균은 약 300 kWh입니다.",
    )

col_ht, col_ck = st.columns(2)
with col_ht:
    heating_type = st.selectbox(
        "현재 주택의 난방 방식",
        list(HEATING_TYPE_MAP.keys()),
        help="현재 사용 중인 난방 연료 방식을 선택해 주세요."
    )
with col_ck:
    cooking_type = st.selectbox(
        "사용하는 취사 기기",
        ["인덕션 (전기)", "도시가스", "LPG"],
    )

# ── 섹션 3: 시뮬레이션 변수 ──
st.markdown('<div class="section-title">3. 시뮬레이션 상수 변수</div>', unsafe_allow_html=True)
col_sim, col_opt = st.columns(2)
with col_sim:
    fuel_inflation = st.slider(
        "가스/등유요금 인상률 (%)", 0.0, 15.0, 11.93, step=0.01,
        help="최근 5년(2021–2025년) 주택용 도시가스 연평균 인상률 11.93%를 기본값으로 설정했습니다. 자유롭게 조정하세요.",
    )
    elec_inflation = st.slider(
        "전기요금 인상률 (%)", 0.0, 15.0, 9.85, step=0.01,
        help="최근 5년(2021–2025년) 주택용 전기 연평균 인상률 9.85%를 기본값으로 설정했습니다. 자유롭게 조정하세요.",
    )
    with st.expander("📑 최근 5년 에너지요금 인상률은? (출처 보기)"):
        st.markdown("""
최근 5년(2021–2025년) 주택용 요금의 연평균 인상률입니다.

- ⚡ 전기: 연평균 9.85%
- 🔥 도시가스: 연평균 11.93%

이 값을 슬라이더 기본값으로 반영했습니다. 향후 전망에 맞게 직접 조정할 수 있습니다.

출처: 에너지통계연보(2021–2024년), 에너지통계월보(2025년)
""")
    solar_install  = st.radio(
        "태양광 설치 여부",
        ["예", "아니오"],
        horizontal=True,
        index=0,
    )
    solar_capa_kw  = st.number_input(
        "태양광 용량 (kW)",
        value=0.0,
        disabled=(solar_install == "아니오"),
        help="태양광 설치 시 발전 용량을 입력해 주세요.",
    )

with col_opt:
    use_subsidy = st.checkbox("보조금 적용 (70%)", value=False)
    st.caption(
        "⚠️ 난방 전기화 사업: 설치비의 70% 지원. "
        "2026년 현재 제주/경남/전남만 지원 가능, 신청 전 확인 필수."
    )

    st.markdown("---")
    st.markdown("**전기 요금제 선택**")
    st.markdown("""
<div class='help-text'>
사용 중인 요금제를 선택해 주세요.
</div>""", unsafe_allow_html=True)

    tariff_choice_simple = st.radio(
        "요금제",
        ["누진제", "일반용", "계시별"],
        label_visibility="collapsed",
        help="누진제: 주택용 저압 누진제 / 일반용: HP 전용 별도 미터 (태양광 영향 없음) / 계시별: 시간대별 요금제",
    )
    if tariff_choice_simple == "일반용" and solar_install == "예":
        st.caption("ℹ️ 일반용은 HP 전용 별도 미터라서 태양광 발전 영향이 없습니다.")
    if tariff_choice_simple == "누진제":
        st.caption(
            "ℹ️ 누진제는 동·하계 월 1,000kWh 초과분에 슈퍼유저요금(736.2원/kWh)이 적용되며, "
            "HP가 일으킨 증분 비용만 난방요금으로 귀속해 계산합니다."
        )

# ── tariff_label 재구성 (기존 TARIFF_LABEL_MAP 5개 키와 호환) ──
# 일반용은 태양광과 무관 — 사용자가 태양광=예 선택해도 미설치로 처리
_apply_solar = (tariff_choice_simple != "일반용") and (solar_install == "예")
if tariff_choice_simple == "일반용":
    tariff_label = "일반용 (HP 전용 미터)"
elif _apply_solar:
    tariff_label = f"{tariff_choice_simple} (태양광 설치)"
else:
    tariff_label = f"{tariff_choice_simple} (태양광 미설치)"

# 입력 변경 감지 → 분석 결과 초기화
_input_signature = (winter_heat_man, winter_elec_kwh, region, house_type, house_size,
                    heating_type, cooking_type, fuel_inflation, elec_inflation,
                    solar_capa_kw, use_subsidy, tariff_label)
if st.session_state.get("_last_input_key") != _input_signature:
    st.session_state.analyzed = False
    st.session_state["_last_input_key"] = _input_signature

if "analyzed" not in st.session_state:
    st.session_state.analyzed = False

if st.button("경제성·환경성 분석 실행", type="primary", use_container_width=True):
    st.session_state.analyzed = True


# ══════════════════════════════════════════════════════════════════════
# 8. 분석 결과
# ══════════════════════════════════════════════════════════════════════

if st.session_state.analyzed:

    # ─── 8-1. 핵심 계산 ───────────────────────────────────────────────
    block_key, tariff_choice, solar_flag = get_block_key(tariff_label, heating_type)
    fuel_key = HEATING_TYPE_MAP[heating_type]   # 이후 모든 곳에서 재사용

    # 광역시도별 월별 난방 비중 가져오기 (Sheet3) — 매핑 실패 시 전국 평균 사용
    sheet3_region_name = REGION_NAME_MAP.get(region, "전국")
    monthly_ratios = REGION_RATIOS.get(sheet3_region_name, REGION_RATIOS.get("전국"))

    # 에너지 사용량 (kWh) — Sheet2 2단계 모델
    # 사용자 연간 난방비 = 1월 입력 ÷ 1월 비중 (지역별)
    user_annual_cost_won = winter_heat_man * 10000 / monthly_ratios[0] if monthly_ratios[0] > 0 else 0
    # 엑셀 Sheet2!C14 공식: 취사가 도시가스/LPG면 0.8475(가스 일부가 취사로 빠짐), 인덕션이면 1.0
    user_heating_share = 0.8475 if cooking_type in ("도시가스", "LPG") else 1.0
    # Sheet2 행 43 — 실제 난방 에너지 사용량 (연료 입력, 효율 적용 전)
    user_fuel_input_kwh = calc_fuel_input_kwh(
        user_annual_cost_won, fuel_key, SHEET2_PARAMS, heating_share=user_heating_share
    )
    # Sheet2 행 44 — 유효 열 수요 (= HP가 만들어야 할 열량, 효율 적용 후)
    user_heat_demand_kwh = calc_heat_demand_kwh(user_fuel_input_kwh, fuel_key, SHEET2_PARAMS)
    # 월별 분배 (행 46-49) + HP 변환 (행 54-57): 사용자 지역 비중 + 사용자 zone 월별 COP
    kwh = calc_kwh_data(user_fuel_input_kwh, user_heat_demand_kwh, monthly_ratios, MONTHLY_COP[zone])

    # ─── [동적 계산] 모든 요금제 — 거창군 묶임 완전 해제 ─────────────
    # 사용자 지역의 실제 HP kWh + 가전 평균 사용량 + 광역시도별 태양광 발전량으로
    # 5개 요금제(누진제 태X/태O, 일반용, 계시별 태X/태O) 청구액 모두 직접 계산.
    # 태양광 설치 시: 사용자 광역시도의 1kW당 발전량 × 설치 용량(kW)
    if solar_flag == "태O" and solar_capa_kw > 0:
        region_solar = SOLAR_KWH.get(REGION_NAME_MAP.get(region, ""), None)
        monthly_solar = [s * solar_capa_kw for s in region_solar] if region_solar else None
    else:
        monthly_solar = None

    # HP 용량 (계산용 숫자 — get_hp_specs는 문자열 "6 kW"라 별도 헬퍼 사용)
    hp_capa_for_calc = get_hp_capacity_kw(house_size)

    # 사용자 입력 1월 전기사용량(kWh)을 월별 비중(Sheet2 행 25)으로 12개월 분배
    # 엑셀 Sheet2!D24:N24 수식과 동일: =$C$24/$C$25*X25
    if ELEC_MONTHLY_RATIOS[0] > 0:
        user_annual_elec_kwh = winter_elec_kwh / ELEC_MONTHLY_RATIOS[0]
        user_monthly_appliance_kwh = [user_annual_elec_kwh * r for r in ELEC_MONTHLY_RATIOS]
    else:
        user_monthly_appliance_kwh = APPLIANCE_KWH  # fallback

    result = calc_dynamic_result(
        tariff_label=tariff_label,
        monthly_hp_kwh=kwh["monthly_hp"],
        monthly_appliance_kwh=user_monthly_appliance_kwh,
        monthly_solar_kwh=monthly_solar,
        hp_capacity_kw=hp_capa_for_calc,
        ex_annual_won=user_annual_cost_won,
    )

    # 보조금 및 투자비
    total_subsidy = SUBSIDY_TOTAL if use_subsidy else 0
    capex_man     = 1000  # 국내 기업 평균 견적 기준 — UI 안내문구 참조
    net_capex_man = max(0, capex_man - total_subsidy)

    # 15년 시뮬레이션
    ann_heat_base = result["ex_annual_man"]
    ann_hp_op     = result["hp_annual_man"]
    years, gas_cum, hp_cum, net_profit, payback_year = simulate_15yr(
        net_capex_man, ann_heat_base, ann_hp_op, fuel_inflation, elec_inflation
    )

    # 월별 기존 난방비 — 사용자 광역시도의 월별 비중으로 안분 (Sheet3 데이터)
    # 비난방월에도 온수 사용분이 있어 0이 되지 않음
    months = list(range(1, 13))
    hdd_zone = HDD_MONTHLY[zone]            # 비고 표시용 (난방월/비난방월 판단)
    jan_ratio = monthly_ratios[0] if monthly_ratios[0] > 0 else 1
    monthly_ex_man = [round(winter_heat_man * monthly_ratios[m-1] / jan_ratio, 2) for m in months]
    # 연간 CO₂ 배출량 — 엑셀 G66~G80 패턴 (유효 열 수요 × HP 배출계수)
    # ⚠️ HP 배출계수는 사용자 기후존의 sCOP에 따라 동적 계산 (엑셀 I10/J10 공식)
    user_scop = SCOP_BY_ZONE[zone]
    ef_hp_2025 = GRID_EF_2025_KGKWH / user_scop   # 엑셀 I10 = 0.4173 / sCOP[zone]
    ef_hp_2038 = GRID_EF_2038_KGKWH / user_scop   # 엑셀 J10 = (83.1/624.5) / sCOP[zone]

    solar_annual_kwh = sum(monthly_solar) if monthly_solar else 0
    co2 = calc_annual_co2_emissions(
        user_annual_cost_won=user_annual_cost_won,
        user_heat_demand_kwh=user_heat_demand_kwh,
        user_heating_share=user_heating_share,
        fuel_key=fuel_key,
        sheet2_params=SHEET2_PARAMS,
        emission_factors_fuel=EMISSION_FACTORS_FUEL,
        emission_factor_hp_2025=ef_hp_2025,
        emission_factor_hp_2038=ef_hp_2038,
        year_idx=0,
    )

    # 월별 CO₂는 연간 배출량을 광역시도 월별 비중으로 안분
    monthly_stats = calc_monthly_stats(
        monthly_ex_man, result["monthly_man"], monthly_ratios,
        co2["ex_kg"], co2["hp_kg"]
    )


    # ─── 8-2. 결과 요약 헤더 ──────────────────────────────────────────
    st.markdown('<div class="section-title">📊 분석 결과 요약</div>', unsafe_allow_html=True)

    # 배지 색상: 태양광 설치=초록, 미설치=노랑, 일반용=파랑
    if solar_flag == "태O":
        badge_cls = "solar-badge-o"
    elif tariff_choice == "일반용":
        badge_cls = "tariff-badge"
    else:
        badge_cls = "solar-badge-x"

    st.markdown(f"""
<div style='margin-bottom:16px;'>
  <span class='{badge_cls}'>{tariff_label}</span>
  <span class='tariff-badge'>난방: {fuel_key}</span>
</div>
""", unsafe_allow_html=True)

    # 핵심 지표 4개
    hp_space, _hp_space_mm, hp_capacity = get_hp_specs(house_size)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("투자 회수 시점", payback_year)
    m2.metric("15년 순이익", f"{net_profit[-1]:,} 만원")
    m3.metric("히트펌프 설치 공간", hp_space)
    m4.metric("적정 히트펌프 용량", hp_capacity)

    # 40평 초과: 단일 기기 한계 초과 → 전문가 상담 안내
    if house_size > 40:
        st.info(
            "ℹ️ 전용면적 40평 초과 주택은 단일 히트펌프로 충분하지 않을 수 있어, "
            "여러 대 설치 또는 맞춤 설계를 위한 **전문가 상담**을 권장합니다. "
            "아래 경제성 수치는 24kW급 1대를 가정한 참고용 추정치입니다."
        )

    # ─── 8-3. 전기요금 분석 ──────────────────────────────────────────
    st.markdown('<div class="section-title">💰 전기요금 분석 (엑셀 데이터 기반)</div>', unsafe_allow_html=True)
    s1, s2, s3 = st.columns(3)
    s1.metric(
        "HP 연간 전기요금",
        f"{result['hp_annual_man']:,.1f} 만원",
        help=f"사용자 1월 입력 → 연간 난방비 → Sheet2 행 43·44 → HP kWh → [{tariff_label}] 단가 적용"
    )
    s2.metric(
        f"기존 연간 난방비 ({fuel_key})",
        f"{result['ex_annual_man']:,.1f} 만원",
        help="사용자 입력한 1월 난방비를 해당 광역시도의 1월 비중으로 나눠 연간 추정"
    )
    s3.metric(
        "연간 절감액",
        f"{result['saving_man']:,.1f} 만원",
        delta=f"{round(result['saving_ratio'] * 100)}% 절감",
    )

    # ─── 8-3-2. 누진제 산정 방식 안내 (슈퍼유저요금 + 증분비용 명시) ──────
    # 누진제 선택 시에만 노출 — HP 전기요금 숫자가 어떻게 나왔는지 투명하게 설명.
    if tariff_choice == "누진제":
        st.markdown(f"""
<div style='margin:8px 0 4px 0; padding:14px 18px; background:#fef7e6;
            border-left:3px solid #d97706; border-radius:8px;'>
  <p style='color:#854d0e; font-size:0.95rem; font-weight:700; margin:0 0 6px 0;'>
    📌 누진제 HP 전기요금은 이렇게 계산했어요
  </p>
  <p style='color:#78350f; font-size:0.9rem; line-height:1.65; margin:0;'>
    위 <b>HP 연간 전기요금</b>에는 히트펌프 전환으로 발생하는 비용이 다음과 같이 반영되어 있습니다.<br>
    ① <b>슈퍼유저요금</b> — 동계(12·1·2월)·하계(7·8월) 월 1,000kWh 초과분은 736.2원/kWh로 가산.
    히트펌프가 겨울 전기 사용량을 1,000kWh 이상으로 끌어올리는 경우가 그대로 반영됩니다.<br>
    ② <b>증분비용 귀속</b> — 히트펌프가 <u>추가로 일으킨 요금</u>(= HP+가전 청구액 − 가전만 청구액)만
    난방요금으로 잡았습니다. 즉 히트펌프 때문에 올라간 누진·슈퍼 구간 비용이 모두 히트펌프 쪽에 귀속됩니다.
  </p>
</div>
""", unsafe_allow_html=True)

    # ─── 8-4. 환경 기여 박스 (Sheet2 행 65 기반) ─────────────────────
    # 사용자 입력 기준 — 현재 사용 중인 난방 연료 vs HP 연간 배출량만 비교.
    # (선택하지 않은 다른 연료는 표시하지 않음 — 연료 교체를 권하는 듯한 오해 방지)
    by_fuel_rows = ""
    _fuel_label_map = {
        "도시가스(콘덴싱)": "도시가스 (콘덴싱)",
        "도시가스(일반)":   "도시가스 (일반)",
        "등유":            "등유",
        "LPG":             "LPG",
    }
    fuel_display = {
        fuel_key:   _fuel_label_map.get(fuel_key, fuel_key),
        "히트펌프": "히트펌프 (전환 시)",
    }
    for f, label in fuel_display.items():
        kg = co2["by_fuel"][f]
        is_current = (f == fuel_key)
        is_hp      = (f == "히트펌프")
        if is_current:
            row_style = "background:#fef7e6; font-weight:700;"
            note = " ← 현재 난방"
        elif is_hp:
            row_style = "background:#e6f4f1; font-weight:700;"
            note = " ← 전환 후"
        else:
            row_style = ""
            note = ""
        by_fuel_rows += (
            f"<tr style='{row_style}'>"
            f"<td style='padding:6px 12px;'>{label}{note}</td>"
            f"<td style='padding:6px 12px; text-align:right;'>{kg:,.0f} kg</td>"
            f"<td style='padding:6px 12px; text-align:right; color:#78716c;'>({kg/1000:.2f} t)</td>"
            f"</tr>"
        )

    st.markdown(f"""
<div class='saving-box'>
  <p class='saving-title'>🌱 우리 가족의 1년 온실가스 배출량</p>
  <p class='saving-sub' style='margin-bottom:12px;'>
    입력하신 정보(1월 난방비 {winter_heat_man}만원, {region}) 기준 추정 결과입니다.
  </p>
  <table style='width:100%; border-collapse:collapse; font-size:0.95rem; color:#064e3b;'>
    <thead>
      <tr style='border-bottom:1px solid #a7d4be;'>
        <th style='padding:6px 12px; text-align:left;'>에너지원</th>
        <th style='padding:6px 12px; text-align:right;'>연간 배출량</th>
        <th style='padding:6px 12px; text-align:right;'>(tCO₂eq)</th>
      </tr>
    </thead>
    <tbody>
      {by_fuel_rows}
    </tbody>
    <tfoot>
      <tr style='border-top:1px solid #a7d4be; font-weight:700;'>
        <td style='padding:8px 12px;'>연간 절감 효과</td>
        <td style='padding:8px 12px; text-align:right;'>{co2['saving_kg']:,.0f} kg</td>
        <td style='padding:8px 12px; text-align:right; color:#064e3b;'>({co2['saving_kg']/1000:.2f} t)</td>
      </tr>
    </tfoot>
  </table>
  <p style='font-size:0.82rem; color:#78716c; margin:10px 0 0 0; line-height:1.5;'>
    * HP 배출량은 난방·온수 등 HP의 전기 사용분 포함 기준 (2026년 그리드).
    15년에 걸쳐 그리드 청정화로 HP 배출량은 점차 감소합니다 (2040년 약 0.20 t).
  </p>
</div>
""", unsafe_allow_html=True)

    # ─── 8-4-2. 에너지 효율 비교 박스 (Sheet2 행 43 + 행 54-57 기반) ─────
    # 기존: 행 43 (실제 사용량) — 사용자가 실제 소비하는 연료 에너지
    # HP:   행 54-57 합 — HP로 전환 시 전기 사용량
    fuel_unit_label = "도시가스" if "도시가스" in fuel_key else fuel_key
    st.markdown(f"""
<div class='saving-box' style='background:var(--bg-teal); border-color:var(--border-teal);'>
  <p class='saving-title' style='color:#134e4a;'>⚡ 우리 가족의 1년 에너지 사용량</p>
  <p class='saving-sub' style='color:#0f766e;'>
    기존 보일러로는 연간 <b>{kwh['fuel_input_annual']:,.0f} kWh</b>의 {fuel_unit_label} 에너지가 필요해요.
    <br>
    히트펌프로 바꾸면 같은 따뜻함을 만드는 데 <b>전기 {kwh['annual_hp']:,.0f} kWh</b>면 충분합니다 —
    🔥 <b>약 {kwh['efficiency']}배 효율</b>로 같은 난방을 1/{kwh['efficiency']:.0f} 에너지로!
  </p>
</div>
""", unsafe_allow_html=True)

    # ─── 8-5. 월별 차트 ──────────────────────────────────────────────
    df_chart = pd.DataFrame({
        "월":               [f"{m}월" for m in months],
        "기존 난방비(만원)": monthly_ex_man,
        "HP 난방요금(만원)": result["monthly_man"],
    }).melt("월", var_name="구분", value_name="금액(만원)")

    chart = alt.Chart(df_chart).mark_bar().encode(
        x=alt.X("월:O", sort=[f"{m}월" for m in months], axis=alt.Axis(labelAngle=0)),
        y=alt.Y("금액(만원):Q"),
        color=alt.Color("구분:N", scale=alt.Scale(
            domain=["기존 난방비(만원)", "HP 난방요금(만원)"],
            range=["#dc2626", "#0d9488"]
        ), legend=alt.Legend(orient="top", title=None)),
        xOffset="구분:N",
        tooltip=["월", "구분", "금액(만원)"],
    ).properties(height=380, title="월별 기존 난방비 vs HP 난방요금")
    st.altair_chart(chart, use_container_width=True)

    # ─── 8-6. 월별 상세 테이블 ────────────────────────────────────────
    with st.expander("📋 월별 상세 데이터"):
        df_detail = pd.DataFrame({
            "월":                [f"{m}월" for m in months],
            "기존 난방비(만원)": monthly_ex_man,
            "HP 난방요금(만원)": result["monthly_man"],
            "월별 절감액(만원)": monthly_stats["savings"],
            "절감률":            monthly_stats["pct"],
            "누적 절감액(만원)": monthly_stats["cumulative"],
            "CO₂ 절감(kg)":     monthly_stats["co2"],
            "HP 전력 사용(kWh)": kwh["monthly_hp"],
        })
        st.dataframe(df_detail, use_container_width=True, hide_index=True)
        st.caption(
            "💡 **HP 난방요금**은 히트펌프의 난방·온수 전기 비용입니다 "
            "(가전·취사 사용분은 제외 — 기존 난방비와 동일 기준 비교)."
        )
        st.caption(
            "🔌 **누진제 산정 방식**: HP가 일으킨 *증분 비용*(= HP+가전 청구액 − 가전만 청구액)으로 계산합니다. "
            "동·하계 월 1,000kWh 초과분에는 슈퍼유저요금(736.2원/kWh)이 반영되어, "
            "히트펌프가 밀어올린 고누진 구간이 난방요금에 정확히 잡힙니다."
        )
        st.caption(
            f"📅 **월별 분배 기준**: 입력하신 광역시도({region})의 평균 월별 난방 비중을 사용자의 1월 입력값에 비례하여 분배한 추정치입니다. "
            "여름철에도 온수·취사 등으로 일부 가스 비용이 발생하는 것이 반영되어 있습니다."
        )

    # ─── 8-7. 장기 차트 ──────────────────────────────────────────────
    st.markdown('<div class="section-title">📈 장기 시뮬레이션</div>', unsafe_allow_html=True)

    g_long_l, g_long_r = st.columns(2)

    # ─── 왼쪽: 15년 총비용 도넛 + 색 카드 ─────────────────────────
    with g_long_l:
        total_ex_15yr = gas_cum[-1]
        total_hp_15yr = hp_cum[-1]

        df_donut = pd.DataFrame({
            "구분": [f"기존 ({fuel_key})", "HP (전기+투자비)"],
            "금액": [total_ex_15yr, total_hp_15yr],
        })

        donut = alt.Chart(df_donut).mark_arc(innerRadius=70, outerRadius=120).encode(
            theta=alt.Theta("금액:Q", stack=True),
            color=alt.Color(
                "구분:N",
                scale=alt.Scale(
                    domain=[f"기존 ({fuel_key})", "HP (전기+투자비)"],
                    range=["#94a3b8", "#2563eb"],
                ),
                legend=None,
            ),
            tooltip=[
                alt.Tooltip("구분:N"),
                alt.Tooltip("금액:Q", format=",.0f", title="15년 누적(만원)"),
            ],
        )

        # 각 조각 위 금액 라벨 (도넛 두께 중간)
        arc_labels = alt.Chart(df_donut).mark_text(
            radius=95, fontSize=15, fontWeight="bold", color="white",
        ).encode(
            theta=alt.Theta("금액:Q", stack=True),
            text=alt.Text("금액:Q", format=",d"),
        )

        st.altair_chart(
            (donut + arc_labels).properties(
                height=340,
                title=alt.TitleParams(
                    text="15년 총비용 비교 (만원)",
                    anchor="start", fontSize=14, color="#1c1917", offset=10,
                ),
            ),
            use_container_width=True,
        )

        # 색깔 매칭 카드 — 회색=기존, 파랑=히트펌프 명시
        st.markdown(f"""
<div style='display:flex; gap:10px; margin-top:4px;'>
  <div style='flex:1; padding:10px 14px; border-radius:8px; border-left:5px solid #94a3b8; background:#f5f5f4;'>
    <div style='font-size:0.82rem; color:#78716c; margin-bottom:4px;'>● 기존 ({fuel_key})</div>
    <div style='font-size:1.15rem; font-weight:700; color:#1c1917;'>{total_ex_15yr:,}만원</div>
  </div>
  <div style='flex:1; padding:10px 14px; border-radius:8px; border-left:5px solid #2563eb; background:#eff6ff;'>
    <div style='font-size:0.82rem; color:#78716c; margin-bottom:4px;'>● 히트펌프 (전기+투자비)</div>
    <div style='font-size:1.15rem; font-weight:700; color:#2563eb;'>{total_hp_15yr:,}만원</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ─── 오른쪽: 연도별 누적 순이익 막대 ──────────────────────────
    with g_long_r:
        # net_profit은 simulate_15yr에서 이미 계산됨:
        #   누적 절감액 - 투자비 → 음수=회수 전, 양수=회수 후
        df_profit = pd.DataFrame({
            "연차": years,
            "누적순이익": net_profit,
            "구분": ["수익" if p >= 0 else "손실" for p in net_profit],
        })

        final_label_df = pd.DataFrame({
            "연차": [years[-1]],
            "누적순이익": [net_profit[-1]],
            "label": [f"{net_profit[-1]:,}만원"],
        })

        bars = alt.Chart(df_profit).mark_bar(cornerRadiusEnd=2).encode(
            x=alt.X("연차:O", title=None,
                    scale=alt.Scale(paddingInner=0.08, paddingOuter=0.05),
                    axis=alt.Axis(
                        labelAngle=0,
                        # 1·5·10·15년차만 라벨 노출 (모바일 가독성)
                        labelExpr=(
                            "(datum.value == 1 || datum.value == 5 || "
                            "datum.value == 10 || datum.value == 15) "
                            "? datum.value + '년차' : ''"
                        ),
                    )),
            y=alt.Y("누적순이익:Q",
                    title="누적 순이익 (만원)",
                    axis=alt.Axis(format=",d")),
            color=alt.Color("구분:N",
                            scale=alt.Scale(domain=["수익", "손실"],
                                            range=["#2563eb", "#7dd3fc"]),
                            legend=alt.Legend(orient="top", title=None,
                                              direction="horizontal")),
            tooltip=[
                alt.Tooltip("연차:O"),
                alt.Tooltip("누적순이익:Q", format=",.0f", title="누적 순이익(만원)"),
                alt.Tooltip("구분:N"),
            ],
        )

        # 0 기준선 (회수 시점 표시)
        zero_rule = alt.Chart(pd.DataFrame({"y": [0]})).mark_rule(
            color="#a8a29e", strokeWidth=1
        ).encode(y="y:Q")

        # 최종 값 라벨
        final_label = alt.Chart(final_label_df).mark_text(
            align="center",
            baseline="bottom" if net_profit[-1] >= 0 else "top",
            dy=-6 if net_profit[-1] >= 0 else 6,
            fontSize=13, fontWeight="bold",
            color="#1c1917",
        ).encode(
            x=alt.X("연차:O"),
            y=alt.Y("누적순이익:Q"),
            text="label:N",
        )

        st.altair_chart(
            (bars + zero_rule + final_label).properties(
                height=420,
                title=alt.TitleParams(
                    text="연도별 누적 순이익 (만원)",
                    anchor="start",
                    fontSize=14,
                    color="#1c1917",
                    offset=10,
                ),
            ),
            use_container_width=True
        )

    # ─── 8-7-2. 환경 효과 (Sheet2 행 65~79 활용 — 2026~2040) ────
    # 그리드 청정화로 HP 배출량이 매년 감소 — 2026년 vs 2040년 효과 비교
    st.markdown('<div class="section-title">🌍 환경 효과 (그리드 청정화 반영)</div>',
                unsafe_allow_html=True)

    # 15년치 배출량 — 엑셀 G66~G80 패턴 (year_idx로 보간 분기)
    # HP 배출계수는 위에서 계산한 동적값 (ef_hp_2025/2038) 그대로 사용
    years_15 = list(range(1, 16))
    co2_15yr_dyn = [
        calc_annual_co2_emissions(
            user_annual_cost_won=user_annual_cost_won,
            user_heat_demand_kwh=user_heat_demand_kwh,
            user_heating_share=user_heating_share,
            fuel_key=fuel_key,
            sheet2_params=SHEET2_PARAMS,
            emission_factors_fuel=EMISSION_FACTORS_FUEL,
            emission_factor_hp_2025=ef_hp_2025,
            emission_factor_hp_2038=ef_hp_2038,
            year_idx=y,
        ) for y in range(15)
    ]
    co2_15_ex = [d["ex_kg"] for d in co2_15yr_dyn]   # 연료는 매년 동일
    co2_15_hp = [d["hp_kg"] for d in co2_15yr_dyn]   # HP는 그리드 청정화로 매년 감소
    co2_15_save = [ex - hp for ex, hp in zip(co2_15_ex, co2_15_hp)]
    co2_15_cum, _cum = [], 0.0
    for s in co2_15_save:
        _cum += s
        co2_15_cum.append(round(_cum, 0))
    total_15yr_kg = co2_15_cum[-1]
    hp_reduction_pct = (1 - co2_15_hp[-1] / co2_15_hp[0]) * 100 if co2_15_hp[0] > 0 else 0

    # 누적 효과 강조 박스
    st.markdown(f"""
<div class='saving-box'>
  <p class='saving-title'>🌍 누적 효과 — 총 {total_15yr_kg:,.0f} kgCO₂ ({total_15yr_kg/1000:.1f} 톤) 감축</p>
  <p class='saving-sub'>
    재생에너지 비중 확대로 그리드(전기) 자체가 청정해지면서 <b>HP 배출량은 매년 감소</b>합니다:
    <br>
    <b>2026년 {co2_15_hp[0]:,.0f} kg → 2040년 {co2_15_hp[-1]:,.0f} kg</b>
    (약 {hp_reduction_pct:.0f}% 감소).
    반면 가스·등유·LPG 보일러는 연소 시점에서 배출하므로 매년 동일합니다.
  </p>
</div>
""", unsafe_allow_html=True)

    # 두 개 차트 — 연간 추이 / 누적 절감
    g3, g4 = st.columns(2)
    with g3:
        st.write("**연간 CO₂ 배출량 추이**")
        df_emit = pd.DataFrame({
            "연차": years_15 + years_15,
            "kg":  co2_15_ex + co2_15_hp,
            "구분": [f"기존 ({fuel_key})"] * 15 + ["HP (그리드 청정화)"] * 15,
        })
        # 매년 두 에너지원 사이 차이를 라벨로 표시 (중간점 위치, ↕ 화살표 + kg)
        # 모바일 가독성을 위해 1·4·7·10·13·15년차만 표시 (3년 간격 + 마지막)
        _label_years = {1, 4, 7, 10, 13, 15}
        df_gap = pd.DataFrame({
            "연차": [y for y in years_15 if y in _label_years],
            "차이(kg)": [co2_15_save[y-1] for y in years_15 if y in _label_years],
            "중간점":  [(co2_15_ex[y-1] + co2_15_hp[y-1]) / 2 for y in years_15 if y in _label_years],
            "차이라벨": [f"↕ {co2_15_save[y-1]:,.0f}" for y in years_15 if y in _label_years],
        })
        line_layer = alt.Chart(df_emit).mark_line(point=True, strokeWidth=2.5).encode(
            x=alt.X("연차:O", title="연차", axis=alt.Axis(labelAngle=0)),
            y=alt.Y("kg:Q", title="연간 배출량 (kg)"),
            color=alt.Color("구분:N", scale=alt.Scale(
                domain=[f"기존 ({fuel_key})", "HP (그리드 청정화)"],
                range=["#dc2626", "#0d9488"]
            ), legend=alt.Legend(orient="top", title=None)),
            tooltip=["연차", "구분", "kg"],
        )
        gap_text = alt.Chart(df_gap).mark_text(
            fontWeight="bold",
            fontSize=11,
            color="#334155",
        ).encode(
            x=alt.X("연차:O"),
            y=alt.Y("중간점:Q"),
            text=alt.Text("차이라벨:N"),
            tooltip=[alt.Tooltip("연차"), alt.Tooltip("차이(kg):Q", format=",.0f")],
        )
        st.altair_chart(
            (line_layer + gap_text).properties(height=380),
            use_container_width=True
        )
    with g4:
        st.write("**누적 CO₂ 절감량**")
        df_cum_co2 = pd.DataFrame({
            "연차": years_15,
            "누적 절감(kg)": co2_15_cum,
        })
        st.altair_chart(
            alt.Chart(df_cum_co2).mark_area(opacity=0.55, color="#10b981").encode(
                x=alt.X("연차:O", title="연차", axis=alt.Axis(labelAngle=0)),
                y=alt.Y("누적 절감(kg):Q"),
                tooltip=["연차", "누적 절감(kg)"],
            ).properties(height=380),
            use_container_width=True
        )

    # ─── 8-8. 면책 안내 ──────────────────────────────────────────────
    st.markdown("---")
    st.caption(
        "⚠️ **본 계산 결과는 간이 추정치로, 실제 결과와 차이가 발생할 수 있습니다.** "
        "표준 가구·평균 단가·평균 기후 데이터를 기반으로 한 시뮬레이션이며, "
        "단열 상태, 사용 습관, 실제 요금제 변동, 설치 환경 등에 따라 결과가 달라질 수 있습니다. "
        "실제 도입 전에는 전문가 견적과 한국전력 공급 약관, 지자체 보조금 공고를 반드시 확인해 주세요."
    )